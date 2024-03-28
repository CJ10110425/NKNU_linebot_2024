import json
import os
import openpyxl
import logging
import time


class Bus:
    def __init__(self):
        logging.info("Initializing BUS Schedules")
        try:
            wb = openpyxl.load_workbook(os.path.abspath(os.path.dirname(__file__)) + "/BUS.xlsx")
            self.Zuoying2YanChao_normal = wb["E04 高鐵左營站-高師大燕巢校區 平日時刻表"]
            self.YanChao2Zuoying_normal = wb["E04 高師大燕巢校區-高鐵左營站 平日時刻表"]
            self.Zuoying2YanChao_weekend = wb["E04 高鐵左營站-高師大燕巢校區 假日時刻表"]
            self.YanChao2Zuoying_weekend = wb["E04 高師大燕巢校區-高鐵左營站 假日時刻表"]
            self.Zuoying2YanChao_vacation = wb["E04 高鐵左營站-高師大燕巢校區 寒暑假時刻表"]
            self.YanChao2Zuoying_vacation = wb["E04 高師大燕巢校區-高鐵左營站 寒暑假時刻表"]
            logging.info("Finished")
        except Exception as e:
            logging.error("An error occurred while reading BUS.xlsx: %s", e)
        self.all_station_Zuoying2YanChao = []
        self.all_station_Yanchao2Zuoying = []
        self.__load_all_station__()

    def __load_all_station__(self):
        for col in self.Zuoying2YanChao_normal['A1': 'S1']:
            for cell in col:
                self.all_station_Zuoying2YanChao.append(cell.value)
        for col in self.YanChao2Zuoying_normal['A1':'S1']:
            for cell in col:
                self.all_station_Yanchao2Zuoying.append(cell.value)

    def __get_current_time__(self):
        return int(time.strftime("%H", time.localtime())), int(time.strftime("%M", time.localtime()))

    def get_next_bus(self, user_station, direction):
        # hour, minute = self.__get_current_time__()
        hour, minute = 10, 40
        station_index = self.all_station_Zuoying2YanChao.index(user_station)
        if direction == "往高師大燕巢校區":
            schedules = {
                "normal": None,
                "weekend": None,
                "vacation": None
            }
            for schedule in [self.Zuoying2YanChao_normal, self.Zuoying2YanChao_weekend, self.Zuoying2YanChao_vacation]:
                row = 2  # 1 是station name
                while True:
                    bus_time = schedule.cell(row, station_index + 1).value
                    bus_time_hour = int(bus_time.hour)
                    bus_time_min = int(bus_time.minute)
                    if bus_time is None or bus_time_hour == hour and bus_time_min >= minute or bus_time_hour > hour:
                        normal_next_bus_schedule_index = row

                        schedule_name = "normal"
                        if schedule == self.Zuoying2YanChao_weekend:
                            schedule_name = "weekend"
                        elif schedule == self.Zuoying2YanChao_vacation:
                            schedule_name = "vacation"
                        schedules[schedule_name] = list(schedule.iter_rows(min_row=row, max_row=row))
                        break
                    row += 1
            return schedules
        elif direction == "往左營高鐵站":
            pass

    def generate_station_selector_flex_message(self, stations):
        flex_message = {"type": "carousel", "contents": []}
        for station in stations:
            flex_message["contents"].append({
                "type": "bubble",
                "size": "micro",
                "hero": {
                    "type": "box",
                    "layout": "vertical",
                    "contents": [
                        {
                            "type": "text",
                            "align": "center",
                            "text": station,
                            "action": {
                                "type": "message",
                                "label": "BUS STATION SELECTOR",
                                "text": f"ʕ •ᴥ•ʔ 燕巢公車 最近的班次 往高師大燕巢校區 我在：{station}"
                            }
                        }
                    ]
                }
            })
        return flex_message


if __name__ == "__main__":
    bus = Bus()
